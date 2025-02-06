import asyncio
from aiogram import Bot, Dispatcher, types
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
from aiogram.filters import Command
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage
import logging
import openpyxl
from datetime import datetime

# Логирование
logging.basicConfig(level=logging.DEBUG)

# Константы
import os
from dotenv import load_dotenv

load_dotenv()  # Загружаем переменные окружения
TOKEN = os.getenv("TOKEN")  # Читаем токен
FILE_PATH = "budjet.xlsx"

# Инициализация бота и диспетчера
bot = Bot(token=TOKEN)
dp = Dispatcher(storage=MemoryStorage())

# Убираем клавиатуру
clear_keyboard = ReplyKeyboardRemove()

# Машина состояний
class IncomeState(StatesGroup):
    amount = State()
    comment = State()

class ExpenseState(StatesGroup):
    category = State()
    subcategory = State()
    amount = State()
    comment = State()

# Клавиатуры
main_menu = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Доход"), KeyboardButton(text="Расход")]
    ],
    resize_keyboard=True
)

more_actions = ReplyKeyboardMarkup(
    keyboard=[[KeyboardButton(text="Да"), KeyboardButton(text="Нет")]],
    resize_keyboard=True
)

expense_categories = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Обязательные траты")],
        [KeyboardButton(text="Продукты домой"), KeyboardButton(text="Еда вне дома")],
        [KeyboardButton(text="Транспорт"), KeyboardButton(text="Товары для дома")],
        [KeyboardButton(text="Товары для себя"), KeyboardButton(text="Медицина")],
        [KeyboardButton(text="Курево"), KeyboardButton(text="Кошки"), KeyboardButton(text="Кофе")],
        [KeyboardButton(text="Развлечения")],
        [KeyboardButton(text="Назад")]
    ],
    resize_keyboard=True
)

mandatory_subcategories = [
    "Коммунальные услуги", "Кредит", "Английский", "Абонентская плата за телефон",
    "Премиум ТГ", "Интернет", "Подписка Яндекс", "Услуги банка", "Кредит"
]

subcategory_keyboard = ReplyKeyboardMarkup(
    keyboard=[[KeyboardButton(text=sub)] for sub in mandatory_subcategories] + [[KeyboardButton(text="Назад")]],
    resize_keyboard=True
)

sum_keyboard = ReplyKeyboardMarkup(
    keyboard=[[KeyboardButton(text="Назад")]],
    resize_keyboard=True
)

# Команда /start
@dp.message(Command("start"))
async def start_command(message: types.Message):
    await message.answer("Мяу! Я твой бот для учета финансов. С чего начнем?", reply_markup=main_menu)

# ДОХОДЫ
@dp.message(lambda message: message.text == "Доход")
async def add_income(message: types.Message, state: FSMContext):
    await message.answer("Введи сумму дохода:", reply_markup=sum_keyboard)
    await state.set_state(IncomeState.amount)

@dp.message(IncomeState.amount)
async def process_income_amount(message: types.Message, state: FSMContext):
    if message.text == "Назад":
        await message.answer("Выбирай:.", reply_markup=main_menu)
        await state.clear()
        return

    if not message.text.isdigit():
        await message.answer("Слушай, бро, я не умею читать копейки и буквы. Введи сумму коректно.")
        return

    await state.update_data(amount=int(message.text))
    await message.answer("Добавь комментарий к доходу:", reply_markup=sum_keyboard)
    await state.set_state(IncomeState.comment)

@dp.message(IncomeState.comment)
async def process_income_comment(message: types.Message, state: FSMContext):
    if message.text == "Назад":
        await message.answer("Введи сумму дохода:", reply_markup=sum_keyboard)
        await state.set_state(IncomeState.amount)
        return  # Выход из функции, чтобы не записывать "Назад" как комментарий

    user_data = await state.get_data()
    amount = user_data["amount"]
    comment = message.text

    try:
        wb = openpyxl.load_workbook(FILE_PATH)
        sheet = wb["Доходы"]
        sheet.append([datetime.today().strftime('%Y-%m-%d'), amount, comment])
        wb.save(FILE_PATH)
        wb.close()
        await message.answer(f"Доход {amount}₽ записан! Нужно внести еще какие-нибудь данные?", reply_markup=more_actions)
        await state.set_state(None)  # Завершаем FSM
    except Exception as e:
        await message.answer("Ошибка записи в таблицу. Веротно она открыта на компе, закрой ее и напиши мне последнее значение.")
        print(f"Ошибка: {e}")

# РАСХОДЫ
@dp.message(lambda message: message.text == "Расход")
async def add_expense(message: types.Message, state: FSMContext):
    await message.answer("Выбери категорию расхода:", reply_markup=expense_categories)
    await state.set_state(ExpenseState.category)

@dp.message(ExpenseState.category)
async def process_expense_category(message: types.Message, state: FSMContext):
    if message.text == "Назад":
        await message.answer("Оке, выбирай:", reply_markup=main_menu)
        await state.clear()
        return

    await state.update_data(category=message.text)

    if message.text == "Обязательные траты":
        await message.answer("Выбери подкатегорию:", reply_markup=subcategory_keyboard)
        await state.set_state(ExpenseState.subcategory)
    else:
        await message.answer("Введи сумму расхода:", reply_markup=sum_keyboard)
        await state.set_state(ExpenseState.amount)

@dp.message(ExpenseState.subcategory)
async def process_expense_subcategory(message: types.Message, state: FSMContext):
    if message.text == "Назад":
        await message.answer("Выбери категорию расхода:", reply_markup=expense_categories)
        await state.set_state(ExpenseState.category)
        return

    await state.update_data(comment=message.text)
    await message.answer("Введи сумму расхода:", reply_markup=sum_keyboard)
    await state.set_state(ExpenseState.amount)

@dp.message(ExpenseState.amount)
async def process_expense_amount(message: types.Message, state: FSMContext):
    if message.text == "Назад":
        user_data = await state.get_data()
        if user_data.get("category") == "Обязательные траты":
            await message.answer("Выбери подкатегорию:", reply_markup=subcategory_keyboard)
            await state.set_state(ExpenseState.subcategory)
        else:
            await message.answer("Выбери категорию расхода:", reply_markup=expense_categories)
            await state.set_state(ExpenseState.category)
        return

    if not message.text.isdigit():
        await message.answer("Слушай, бро, я не умею читать копейки и буквы. Введи сумму корректно.")
        return

    amount = int(message.text)  # Исправлено: теперь amount получает значение
    await state.update_data(amount=amount)

    user_data = await state.get_data()
    if user_data["category"] == "Обязательные траты":
        comment = user_data["comment"]  # Берем подкатегорию как комментарий
    else:
        await message.answer("Добавь комментарий к расходу:", reply_markup=sum_keyboard)
        await state.set_state(ExpenseState.comment)
        return

    # Если категория "Обязательные траты", сразу записываем в таблицу
    try:
        wb = openpyxl.load_workbook(FILE_PATH)
        sheet = wb["Расходы"]
        sheet.append([datetime.today().strftime('%Y-%m-%d'), user_data["category"], amount, comment])
        wb.save(FILE_PATH)
        wb.close()
        await message.answer(f"Расход {amount}₽ записан! Нужно внести еще какие-нибудь данные?", reply_markup=more_actions)
        await state.clear()
    except Exception as e:
        await message.answer("Ошибка записи в таблицу. Вероятно, она открыта на компе, закрой её и напиши мне последнее значение.")
        print(f"Ошибка: {e}")
    

@dp.message(ExpenseState.comment)
async def process_expense_comment(message: types.Message, state: FSMContext):
    if message.text == "Назад":
        await message.answer("Введи сумму расхода:", reply_markup=sum_keyboard)
        await state.set_state(ExpenseState.amount)
        return  # Выход из функции, чтобы не записывать "Назад" как комментарий

    user_data = await state.get_data()
    category = user_data["category"]
    amount = user_data["amount"]
    comment = message.text

    try:
        wb = openpyxl.load_workbook(FILE_PATH)
        sheet = wb["Расходы"]
        sheet.append([datetime.today().strftime('%Y-%m-%d'), category, amount, comment])
        wb.save(FILE_PATH)
        wb.close()
        await message.answer(f"Расход {amount}₽ записан! Нужно внести еще какие-нибудь данные?", reply_markup=more_actions)
        await state.set_state(None)  # Завершаем FSM
    except Exception as e:
        await message.answer("Ошибка записи в таблицу. Веротно она открыта на компе, закрой ее и напиши мне последнее значение.")
        print(f"Ошибка: {e}")

@dp.message(lambda message: message.text in ["Да", "Нет"])
async def process_more_actions(message: types.Message):
    if message.text == "Да":
        await message.answer("Оке, выбирай:", reply_markup=main_menu)
    else:
        await message.answer("Если что, я всегда рядом!", reply_markup=main_menu)

# --- Запуск бота ---
async def main():
    print("Запуск бота...")
    try:
        me = await bot.get_me()
        print(f"Бот запущен: {me.username}")
        await dp.start_polling(bot)
    except Exception as e:
        print(f"Ошибка: {e}")

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("Бот остановлен вручную")
    except Exception as e:
        print(f"Ошибка: {e}")